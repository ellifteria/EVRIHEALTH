import openpyxl

def get_chargemaster(hospital):
    path = "static/chargemaster/{}.xlsx".format(hospital)
    
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    charges = {}
    
    for i in range(1, m_row + 1):
        charges[sheet_obj.cell(row = i, column = 2).value] = sheet_obj.cell(row = i, column = 3).value
    return charges
