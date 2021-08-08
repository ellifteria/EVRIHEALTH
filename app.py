from flask import Flask, render_template, request, redirect, session
import datetime
import random
import string
from PyPDF2 import PdfFileReader, PdfFileWriter
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas
import io
import os
import openpyxl

def get_chargemaster(hospital):
    path = "static/chargemaster/{}.xlsx".format(hospital)
    
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    charges = {}
    
    for i in range(1, m_row + 1):
        charges[sheet_obj.cell(row = i, column = 2).value] = sheet_obj.cell(row = i, column = 3).value
    return charges

federal_poverty_guidelines = {
    1: 12880,
    2: 17420,
    3: 21960,
    4: 26500,
    5: 31040,
    6: 35580,
    7: 40120,
    8: 44660
}

def calculate_threshold(household_size):
    threshold_extension = 0
    while household_size > 8:
        household_size -= 1
        threshold_extension += 4540
    return federal_poverty_guidelines[household_size] + threshold_extension

def determine_financially_indignant(household_size, estimated_annual_income):
    poverty_threshold = calculate_threshold(household_size)
    if estimated_annual_income <= poverty_threshold * 2:
        return 1.00
    return 0

def determine_medically_indignant(household_size, estimated_annual_income, balance_due):
    poverty_threshold = calculate_threshold(household_size)
    if estimated_annual_income <= poverty_threshold * 2.5:
        if balance_due / estimated_annual_income > 0.05:
            return 0.95
    elif estimated_annual_income <= poverty_threshold * 3:
        if balance_due / estimated_annual_income > 0.05:
            return 0.90
    elif estimated_annual_income <= poverty_threshold * 3.5:
        if balance_due / estimated_annual_income > 0.05:
            return 0.85
    elif estimated_annual_income <= poverty_threshold * 4:
        if balance_due / estimated_annual_income > 0.10:
            return 0.80
    elif estimated_annual_income <= poverty_threshold * 5:
        if balance_due / estimated_annual_income > 0.10:
            return 0.75
    return 0

def determine_tier2_medically_indignant(household_size, estimated_annual_income, balance_due):
    if estimated_annual_income > calculate_threshold(household_size) * 5:
        if balance_due >= estimated_annual_income:
            return 0.95
        elif balance_due >= estimated_annual_income * 0.8:
            return 0.90
        elif balance_due >= estimated_annual_income * 0.6:
            return 0.85
        elif balance_due >= estimated_annual_income * 0.4:
            return 0.80
        elif balance_due >= estimated_annual_income * 0.2:
            return 0.75
    return 0

def determine_catastrophic_medically_indignant(household_size, estimated_annual_income, balance_due):
    if balance_due > estimated_annual_income:
        if estimated_annual_income <= calculate_threshold(household_size) * 5:
            return 0.975
        else:
            return 0.95
    return 0



def create_pdf(destination, template, data_to_add):
    template_pdf = PdfFileReader(open(template, "rb"))
    output = PdfFileWriter()

    packet = io.BytesIO()

    pdf_canvas = Canvas(packet, pagesize=LETTER)
    for page, text_to_add in data_to_add.items():
        if text_to_add is not None:
            for data in text_to_add:
                pdf_canvas.drawString(data[1], data[2], data[0])
        pdf_canvas.showPage()
    pdf_canvas.save()

    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    for page_number, text_to_add in data_to_add.items():
        page = template_pdf.getPage(page_number)
        page.mergePage(new_pdf.getPage(page_number))
        page.compressContentStreams()
        output.addPage(page)

    outputStream = open(destination, "wb")
    output.write(outputStream)
    outputStream.close()

def delete_file(target):
    if os.path.exists(target):
        os.remove(target)

def fill_out_THP_application(
        pdf_id, curr_date, guarantor_name, last_name, first_name, MI, DOS,
        hospital_acct, medical_rcrd, facility, social_security, DOB, 
        marital_status, minors, living_with, legal_minor, patient_employed,
        spouse_employed, med_insurance, disability, disability_length,
        veteran, spouse_name, child1_name, child1_age, child2_name, child2_age,
        child3_name, child3_age, child4_name, child4_age, patient_gross,
        patient_net, spouse_gross, spouse_net, dependants_gross,
        dependants_net, pub_asst_gross, pub_asst_net, food_stamps_gross,
        food_stamps_net, soc_serc_gross, soc_serc_net, unemp_gross, unemp_net,
        strk_ben_gross, strk_ben_net, work_comp_gross, work_comp_net,
        alim_gross, alim_net, chld_sup_gross, chld_sup_net, mil_all_gross,
        mil_all_net, pen_gross, pen_net, inc_gross, inc_net, rent, utilities,
        car, groceries, credit, other_descr, other, checking, saving, CDs_IRAs,
        other_invst, properties, employer_name, spouse_employer_name,
        employer_phone, spouse_employer_phone, employer_address, spouse_employer_address,
        occupation, spouse_occupation, medicaid, county, donate, liability, assist,
        assist_identity, assist_amnt, other_info, lost_earnings, lost_time
    ):

    opt_facility = {
        "Texas Health Center for Diagnostics & Surgery Plano" : [1, 3.4],
        "Texas Health Harris Methodist Southlake": [3.3, 3.4],
        "Texas Health Presbyterian Hospital Flower Mound": [5.75, 3.4],
        "Texas Health Presbyterian Hospital Rockwall": [1, 3.9],
        "Texas Institute for Surgery at Texas Health Presbyterian Dallas": [3.3, 3.9]
    }
    opt_marital_status = {
        "Married": [1.7, 3],
        "Single": [3, 3],
        "Divorced": [4.3, 3],
        "Widowed": [5.65, 3],
        "Separated": [7.15, 3]
    }
    opt_minors = {
        "Yes": [4.6, 3.38],
        "No": [5.6, 3.38]
    }
    opt_living_with = {
        "Yes": [4.6, 3.55],
        "No": [5.6, 3.55]
    }
    opt_legal_minor = {
        "Yes": [4.6, 3.72],
        "No": [5.6, 3.72]
    }
    opt_patient_employed = {
        "Yes": [4.6, 3.89],
        "No": [5.6, 3.89]
    }
    opt_spouse_employed = {
        "Yes": [4.6, 4.06],
        "No": [5.6, 4.06]
    }
    opt_med_insurance = {
        "Yes": [4.6, 4.23],
        "No": [5.6, 4.23]
    }
    opt_disability = {
        "Yes": [4.6, 4.4],
        "No": [5.6, 4.4]
    }
    opt_veteran = {
        "Yes": [4.6, 4.57],
        "No": [5.6, 4.57]
    }

    total_gross = int(inc_gross) + int(pen_gross) + int(alim_gross) + int(unemp_gross) + int(food_stamps_gross) + int(dependants_gross)
    total_gross += int(work_comp_gross) + int(strk_ben_gross) + int(soc_serc_gross) + int(pub_asst_gross) + int(chld_sup_gross)
    total_gross += int(patient_gross) + int(mil_all_gross) + int(spouse_gross)
    total_net = int(inc_net) + int(pen_net) + int(alim_net) + int(unemp_net) + int(food_stamps_net) + int(dependants_net)
    total_net += int(work_comp_net) + int(strk_ben_net) + int(soc_serc_net) + int(pub_asst_net) + int(chld_sup_net)
    total_net += int(patient_net) + int(mil_all_net) + int(spouse_net)
    total = int(rent) + int(utilities) + int(car) + int(groceries) + int(credit) + int(other)

    opt_medicaid = {
        "Yes": [5.95, 3.4],
        "No": [6.59, 3.4]
    }
    opt_county = {
        "Yes": [5.95, 3.57],
        "No": [6.59, 3.57]
    }
    opt_donate = {
        "Yes": [5.95, 3.72],
        "No": [6.59, 3.72]
    }
    opt_liability = {
        "Yes": [5.95, 4.08],
        "No": [6.59, 4.08]
    }
    opt_assist = {
        "Yes": [5.95, 4.24],
        "No": [6.59, 4.24]
    }

    patient_data = {
        0: [
            [curr_date, 1.5*inch, LETTER[1] - 2.1*inch],
            [guarantor_name, 5.3*inch, LETTER[1] - 2.1*inch],
            [first_name + " " + MI + " " + last_name, 2.1*inch, LETTER[1] - 2.4*inch],
            [DOS, 5.1*inch, LETTER[1] - 2.4*inch],
            [hospital_acct, 2.35*inch, LETTER[1] - 2.7*inch],
            [medical_rcrd, 5.25*inch, LETTER[1] - 2.7*inch],
            ["X", opt_facility[facility][0]*inch, LETTER[1] - opt_facility[facility][1]*inch]
        ],
        1: [
            [last_name, 2.4*inch, LETTER[1] - 2.375*inch],
            [first_name, 5*inch, LETTER[1] - 2.375*inch],
            [MI, 7.2*inch, LETTER[1] - 2.375*inch],
            [social_security, 2.1*inch, LETTER[1] - 2.7*inch],
            [DOB, 3.85*inch, LETTER[1] - 2.7*inch],
            [hospital_acct, 6.1*inch, LETTER[1] - 2.7*inch],
            
            ["X", opt_marital_status[marital_status][0]*inch, LETTER[1] - opt_marital_status[marital_status][1]*inch],
            ["X", opt_minors[minors][0]*inch, LETTER[1] - opt_minors[minors][1]*inch],
            ["X", opt_living_with[living_with][0]*inch, LETTER[1] - opt_living_with[living_with][1]*inch],
            ["X", opt_legal_minor[legal_minor][0]*inch, LETTER[1] - opt_legal_minor[legal_minor][1]*inch],
            ["X", opt_patient_employed[patient_employed][0]*inch, LETTER[1] - opt_patient_employed[patient_employed][1]*inch],
            ["X", opt_spouse_employed[spouse_employed][0]*inch, LETTER[1] - opt_spouse_employed[spouse_employed][1]*inch],
            ["X", opt_med_insurance[med_insurance][0]*inch, LETTER[1] - opt_med_insurance[med_insurance][1]*inch],
            ["X", opt_disability[disability][0]*inch, LETTER[1] - opt_disability[disability][1]*inch],
            [disability_length, 3.15*inch, LETTER[1] - 4.4*inch],
            ["X", opt_veteran[veteran][0]*inch, LETTER[1] - opt_veteran[veteran][1]*inch],

            [spouse_name, 1.65*inch, LETTER[1] - 5*inch],
            [child1_name, 1.4*inch, LETTER[1] - 5.2*inch],
            [child1_age, 3.75*inch, LETTER[1] - 5.2*inch],
            [child2_name, 1.4*inch, LETTER[1] - 5.37*inch],
            [child2_age, 3.75*inch, LETTER[1] - 5.37*inch],
            [child3_name, 1.4*inch, LETTER[1] - 5.54*inch],
            [child3_age, 3.75*inch, LETTER[1] - 5.54*inch],
            [child4_name, 1.4*inch, LETTER[1] - 5.71*inch],
            [child4_age, 3.75*inch, LETTER[1] - 5.71*inch],

            [patient_gross, 2.6*inch, LETTER[1] - 6.32*inch],
            [patient_net, 3.7*inch, LETTER[1] - 6.32*inch],
            [spouse_gross, 2.6*inch, LETTER[1] - 6.49*inch],
            [spouse_net, 3.7*inch, LETTER[1] - 6.49*inch],
            [dependants_gross, 2.6*inch, LETTER[1] - 6.66*inch],
            [dependants_net, 3.7*inch, LETTER[1] - 6.66*inch],
            [pub_asst_gross, 2.6*inch, LETTER[1] - 6.83*inch],
            [pub_asst_net, 3.7*inch, LETTER[1] - 6.83*inch],
            [food_stamps_gross, 2.6*inch, LETTER[1] - 7*inch],
            [food_stamps_net, 3.7*inch, LETTER[1] - 7*inch],
            [soc_serc_gross, 2.6*inch, LETTER[1] - 7.17*inch],
            [soc_serc_net, 3.7*inch, LETTER[1] - 7.17*inch],
            [unemp_gross, 2.6*inch, LETTER[1] - 7.34*inch],
            [unemp_net, 3.7*inch, LETTER[1] - 7.34*inch],
            [strk_ben_gross, 2.6*inch, LETTER[1] - 7.51*inch],
            [strk_ben_net, 3.7*inch, LETTER[1] - 7.51*inch],
            [work_comp_gross, 2.6*inch, LETTER[1] - 7.81*inch],
            [work_comp_net, 3.7*inch, LETTER[1] - 7.81*inch],
            [alim_gross, 2.6*inch, LETTER[1] - 7.98*inch],
            [alim_net, 3.7*inch, LETTER[1] - 7.98*inch],
            [chld_sup_gross, 2.6*inch, LETTER[1] - 8.15*inch],
            [chld_sup_net, 3.7*inch, LETTER[1] - 8.15*inch],
            [mil_all_gross, 2.6*inch, LETTER[1] - 8.32*inch],
            [mil_all_net, 3.7*inch, LETTER[1] - 8.32*inch],
            [pen_gross, 2.6*inch, LETTER[1] - 8.49*inch],
            [pen_net, 3.7*inch, LETTER[1] - 8.49*inch],
            [inc_gross, 2.6*inch, LETTER[1] - 8.97*inch],
            [inc_net, 3.7*inch, LETTER[1] - 8.97*inch],
            [str(total_gross), 2.6*inch, LETTER[1] - 9.23*inch],
            [str(total_net), 3.7*inch, LETTER[1] - 9.23*inch],

            [rent, 6.91*inch, LETTER[1] - 6.32*inch],
            [utilities, 6.91*inch, LETTER[1] - 6.49*inch],
            [car, 6.91*inch, LETTER[1] - 6.66*inch],
            [groceries, 6.91*inch, LETTER[1] - 6.83*inch],
            [credit, 6.91*inch, LETTER[1] - 7*inch],
            [other_descr, 5.03*inch, LETTER[1] - 7.33*inch],
            [other, 6.91*inch, LETTER[1] - 7.33*inch],
            [str(total), 6.91*inch, LETTER[1] - 7.82*inch],

            [checking, 4.19*inch, LETTER[1] - 9.67*inch],
            [saving, 4.19*inch, LETTER[1] - 9.84*inch],
            [CDs_IRAs, 4.19*inch, LETTER[1] - 10.01*inch],
            [other_invst, 4.19*inch, LETTER[1] - 10.18*inch],
            [properties, 4.19*inch, LETTER[1] - 10.35*inch]
        ],
        2: [
            [employer_name, 2.2*inch, LETTER[1] - 2.53*inch],
            [spouse_employer_name, 5.8*inch, LETTER[1] - 2.53*inch],
            [employer_phone, 2.2*inch, LETTER[1] - 2.72*inch],
            [spouse_employer_phone, 5.8*inch, LETTER[1] - 2.72*inch],
            [employer_address, 2.2*inch, LETTER[1] - 2.91*inch],
            [spouse_employer_address, 5.8*inch, LETTER[1] - 2.91*inch],
            [occupation, 2.2*inch, LETTER[1] - 3.10*inch],
            [spouse_occupation, 5.8*inch, LETTER[1] - 3.10*inch],

            ["X", opt_medicaid[medicaid][0]*inch, LETTER[1] - opt_medicaid[medicaid][1]*inch],
            ["X", opt_county[county][0]*inch, LETTER[1] - opt_county[county][1]*inch],
            ["X", opt_donate[donate][0]*inch, LETTER[1] - opt_donate[donate][1]*inch],
            ["X", opt_liability[liability][0]*inch, LETTER[1] - opt_liability[liability][1]*inch],
            ["X", opt_assist[assist][0]*inch, LETTER[1] - opt_assist[assist][1]*inch],
            [assist_identity, 5.8*inch, LETTER[1] - 4.4*inch],
            [assist_amnt, 5.8*inch, LETTER[1] - 4.55*inch],

            [other_info, 0.85*inch, LETTER[1] - 5.17*inch],

            [lost_earnings, 5.99*inch, LETTER[1] - 5.92*inch],
            [lost_time, 5.9*inch, LETTER[1] - 6.2*inch]
        ]
    }
    create_pdf(pdf_id, "static/THP Website Charity Application English.pdf", patient_data)

def fill_out_charitycare_application(
        pdf_id, curr_date, guarantor_name, last_name, first_name, MI, DOS,
        hospital_acct, medical_rcrd, facility, social_security, DOB, 
        marital_status, minors, living_with, legal_minor, patient_employed,
        spouse_employed, med_insurance, disability, disability_length,
        veteran, spouse_name, child1_name, child1_age, child2_name, child2_age,
        child3_name, child3_age, child4_name, child4_age, patient_gross,
        patient_net, spouse_gross, spouse_net, dependants_gross,
        dependants_net, pub_asst_gross, pub_asst_net, food_stamps_gross,
        food_stamps_net, soc_serc_gross, soc_serc_net, unemp_gross, unemp_net,
        strk_ben_gross, strk_ben_net, work_comp_gross, work_comp_net,
        alim_gross, alim_net, chld_sup_gross, chld_sup_net, mil_all_gross,
        mil_all_net, pen_gross, pen_net, inc_gross, inc_net, rent, utilities,
        car, groceries, credit, other_descr, other, checking, saving, CDs_IRAs,
        other_invst, properties, employer_name, spouse_employer_name,
        employer_phone, spouse_employer_phone, employer_address, spouse_employer_address,
        occupation, spouse_occupation, medicaid, county, donate, liability, assist,
        assist_identity, assist_amnt, other_info, lost_earnings, lost_time
    ):

    opt_facility = {
        "Texas Health Allen" : [1.05, 3.16],
        "Texas Health Denton": [3.29, 3.16],
        "Texas Health Prosper": [5.57, 3.16],
        "Texas Health Alliance": [1.05, 3.45],
        "Texas Health Frisco": [3.29, 3.45],
        "Texas Health Recovery & Wellness Center": [5.57, 3.45],
        "Texas Health Arlington Memorial": [1.05, 3.76],
        "Texas Health Fort Worth": [3.29, 3.76],
        "Texas Health Southwest Fort Worth": [5.57, 3.76],
        "Texas Health Azle": [1.05, 4.08],
        "Texas Health Heart & Vascular Hospital Arlington": [3.29, 4.08],
        "Texas Health Specialty Hospital": [5.57, 4.08],
        "Texas Health Burleson": [1.05, 4.42],
        "Texas Health HEB": [3.29, 4.42],
        "Texas Health Springwood": [5.57, 4.42],
        "Texas Health Cleburne": [1.05, 4.77],
        "Texas Health Kaufman": [3.29, 4.77],
        "Texas Health Stephenville": [5.57, 4.77],
        "Texas Health Dallas": [1.05, 5.16],
        "Texas Health Plano": [3.29, 5.16],
        "Texas Health Urgent Care": [5.57, 5.16]
    }
    opt_marital_status = {
        "Married": [1.7, 3],
        "Single": [3, 3],
        "Divorced": [4.3, 3],
        "Widowed": [5.65, 3],
        "Separated": [7.15, 3]
    }
    opt_minors = {
        "Yes": [4.6, 3.38],
        "No": [5.6, 3.38]
    }
    opt_living_with = {
        "Yes": [4.6, 3.55],
        "No": [5.6, 3.55]
    }
    opt_legal_minor = {
        "Yes": [4.6, 3.72],
        "No": [5.6, 3.72]
    }
    opt_patient_employed = {
        "Yes": [4.6, 3.89],
        "No": [5.6, 3.89]
    }
    opt_spouse_employed = {
        "Yes": [4.6, 4.06],
        "No": [5.6, 4.06]
    }
    opt_med_insurance = {
        "Yes": [4.6, 4.23],
        "No": [5.6, 4.23]
    }
    opt_disability = {
        "Yes": [4.6, 4.4],
        "No": [5.6, 4.4]
    }
    opt_veteran = {
        "Yes": [4.6, 4.57],
        "No": [5.6, 4.57]
    }

    total_gross = int(inc_gross) + int(pen_gross) + int(alim_gross) + int(unemp_gross) + int(food_stamps_gross) + int(dependants_gross)
    total_gross += int(work_comp_gross) + int(strk_ben_gross) + int(soc_serc_gross) + int(pub_asst_gross) + int(chld_sup_gross)
    total_gross += int(patient_gross) + int(mil_all_gross) + int(spouse_gross)
    total_net = int(inc_net) + int(pen_net) + int(alim_net) + int(unemp_net) + int(food_stamps_net) + int(dependants_net)
    total_net += int(work_comp_net) + int(strk_ben_net) + int(soc_serc_net) + int(pub_asst_net) + int(chld_sup_net)
    total_net += int(patient_net) + int(mil_all_net) + int(spouse_net)
    total = int(rent) + int(utilities) + int(car) + int(groceries) + int(credit) + int(other)

    opt_medicaid = {
        "Yes": [5.95, 3.4],
        "No": [6.59, 3.4]
    }
    opt_county = {
        "Yes": [5.95, 3.57],
        "No": [6.59, 3.57]
    }
    opt_donate = {
        "Yes": [5.95, 3.72],
        "No": [6.59, 3.72]
    }
    opt_liability = {
        "Yes": [5.95, 4.08],
        "No": [6.59, 4.08]
    }
    opt_assist = {
        "Yes": [5.95, 4.24],
        "No": [6.59, 4.24]
    }

    patient_data = {
        0: [
            [curr_date, 1.5*inch, LETTER[1] - 2.1*inch],
            [guarantor_name, 5.3*inch, LETTER[1] - 2.1*inch],
            [first_name + " " + MI + " " + last_name, 2.1*inch, LETTER[1] - 2.44*inch],
            [DOS, 6.27*inch, LETTER[1] - 2.44*inch],
            [hospital_acct, 2.35*inch, LETTER[1] - 2.76*inch],
            [medical_rcrd, 6.41*inch, LETTER[1] - 2.76*inch],
            ["X", opt_facility[facility][0]*inch, LETTER[1] - opt_facility[facility][1]*inch]
        ],
        1: [
            [last_name, 2.4*inch, LETTER[1] - 2.375*inch],
            [first_name, 5*inch, LETTER[1] - 2.375*inch],
            [MI, 7.2*inch, LETTER[1] - 2.375*inch],
            [social_security, 2.1*inch, LETTER[1] - 2.7*inch],
            [DOB, 3.85*inch, LETTER[1] - 2.7*inch],
            [hospital_acct, 6.1*inch, LETTER[1] - 2.7*inch],
            
            ["X", opt_marital_status[marital_status][0]*inch, LETTER[1] - opt_marital_status[marital_status][1]*inch],
            ["X", opt_minors[minors][0]*inch, LETTER[1] - opt_minors[minors][1]*inch],
            ["X", opt_living_with[living_with][0]*inch, LETTER[1] - opt_living_with[living_with][1]*inch],
            ["X", opt_legal_minor[legal_minor][0]*inch, LETTER[1] - opt_legal_minor[legal_minor][1]*inch],
            ["X", opt_patient_employed[patient_employed][0]*inch, LETTER[1] - opt_patient_employed[patient_employed][1]*inch],
            ["X", opt_spouse_employed[spouse_employed][0]*inch, LETTER[1] - opt_spouse_employed[spouse_employed][1]*inch],
            ["X", opt_med_insurance[med_insurance][0]*inch, LETTER[1] - opt_med_insurance[med_insurance][1]*inch],
            ["X", opt_disability[disability][0]*inch, LETTER[1] - opt_disability[disability][1]*inch],
            [disability_length, 3.15*inch, LETTER[1] - 4.4*inch],
            ["X", opt_veteran[veteran][0]*inch, LETTER[1] - opt_veteran[veteran][1]*inch],

            [spouse_name, 1.65*inch, LETTER[1] - 5*inch],
            [child1_name, 1.4*inch, LETTER[1] - 5.2*inch],
            [child1_age, 3.75*inch, LETTER[1] - 5.2*inch],
            [child2_name, 1.4*inch, LETTER[1] - 5.37*inch],
            [child2_age, 3.75*inch, LETTER[1] - 5.37*inch],
            [child3_name, 1.4*inch, LETTER[1] - 5.54*inch],
            [child3_age, 3.75*inch, LETTER[1] - 5.54*inch],
            [child4_name, 1.4*inch, LETTER[1] - 5.71*inch],
            [child4_age, 3.75*inch, LETTER[1] - 5.71*inch],

            [patient_gross, 2.62*inch, LETTER[1] - 6.32*inch],
            [patient_net, 3.76*inch, LETTER[1] - 6.32*inch],
            [spouse_gross, 2.62*inch, LETTER[1] - 6.49*inch],
            [spouse_net, 3.76*inch, LETTER[1] - 6.49*inch],
            [dependants_gross, 2.62*inch, LETTER[1] - 6.66*inch],
            [dependants_net, 3.76*inch, LETTER[1] - 6.66*inch],
            [pub_asst_gross, 2.62*inch, LETTER[1] - 6.83*inch],
            [pub_asst_net, 3.76*inch, LETTER[1] - 6.83*inch],
            [food_stamps_gross, 2.62*inch, LETTER[1] - 7*inch],
            [food_stamps_net, 3.76*inch, LETTER[1] - 7*inch],
            [soc_serc_gross, 2.62*inch, LETTER[1] - 7.17*inch],
            [soc_serc_net, 3.76*inch, LETTER[1] - 7.17*inch],
            [unemp_gross, 2.62*inch, LETTER[1] - 7.34*inch],
            [unemp_net, 3.76*inch, LETTER[1] - 7.34*inch],
            [strk_ben_gross, 2.62*inch, LETTER[1] - 7.51*inch],
            [strk_ben_net, 3.76*inch, LETTER[1] - 7.51*inch],
            [work_comp_gross, 2.62*inch, LETTER[1] - 7.81*inch],
            [work_comp_net, 3.76*inch, LETTER[1] - 7.81*inch],
            [alim_gross, 2.62*inch, LETTER[1] - 7.98*inch],
            [alim_net, 3.76*inch, LETTER[1] - 7.98*inch],
            [chld_sup_gross, 2.62*inch, LETTER[1] - 8.15*inch],
            [chld_sup_net, 3.76*inch, LETTER[1] - 8.15*inch],
            [mil_all_gross, 2.62*inch, LETTER[1] - 8.32*inch],
            [mil_all_net, 3.76*inch, LETTER[1] - 8.32*inch],
            [pen_gross, 2.62*inch, LETTER[1] - 8.49*inch],
            [pen_net, 3.76*inch, LETTER[1] - 8.49*inch],
            [inc_gross, 2.62*inch, LETTER[1] - 8.97*inch],
            [inc_net, 3.76*inch, LETTER[1] - 8.97*inch],
            [str(total_gross), 2.62*inch, LETTER[1] - 9.23*inch],
            [str(total_net), 3.76*inch, LETTER[1] - 9.23*inch],

            [rent, 6.91*inch, LETTER[1] - 6.32*inch],
            [utilities, 6.91*inch, LETTER[1] - 6.49*inch],
            [car, 6.91*inch, LETTER[1] - 6.66*inch],
            [groceries, 6.91*inch, LETTER[1] - 6.83*inch],
            [credit, 6.91*inch, LETTER[1] - 7*inch],
            [other_descr, 5.07*inch, LETTER[1] - 7.33*inch],
            [other, 6.91*inch, LETTER[1] - 7.33*inch],
            [str(total), 6.91*inch, LETTER[1] - 7.81*inch],

            [checking, 4.19*inch, LETTER[1] - 9.67*inch],
            [saving, 4.19*inch, LETTER[1] - 9.84*inch],
            [CDs_IRAs, 4.19*inch, LETTER[1] - 10.01*inch],
            [other_invst, 4.19*inch, LETTER[1] - 10.18*inch],
            [properties, 4.19*inch, LETTER[1] - 10.35*inch]
        ],
        2: [
            [employer_name, 2.2*inch, LETTER[1] - 2.53*inch],
            [spouse_employer_name, 5.8*inch, LETTER[1] - 2.53*inch],
            [employer_phone, 2.2*inch, LETTER[1] - 2.72*inch],
            [spouse_employer_phone, 5.8*inch, LETTER[1] - 2.72*inch],
            [employer_address, 2.2*inch, LETTER[1] - 2.91*inch],
            [spouse_employer_address, 5.8*inch, LETTER[1] - 2.91*inch],
            [occupation, 2.2*inch, LETTER[1] - 3.10*inch],
            [spouse_occupation, 5.8*inch, LETTER[1] - 3.10*inch],

            ["X", opt_medicaid[medicaid][0]*inch, LETTER[1] - opt_medicaid[medicaid][1]*inch],
            ["X", opt_county[county][0]*inch, LETTER[1] - opt_county[county][1]*inch],
            ["X", opt_donate[donate][0]*inch, LETTER[1] - opt_donate[donate][1]*inch],
            ["X", opt_liability[liability][0]*inch, LETTER[1] - opt_liability[liability][1]*inch],
            ["X", opt_assist[assist][0]*inch, LETTER[1] - opt_assist[assist][1]*inch],
            [assist_identity, 5.8*inch, LETTER[1] - 4.4*inch],
            [assist_amnt, 5.8*inch, LETTER[1] - 4.55*inch],

            [other_info, 0.85*inch, LETTER[1] - 5.17*inch],

            [lost_earnings, 5.99*inch, LETTER[1] - 5.92*inch],
            [lost_time, 5.9*inch, LETTER[1] - 6.2*inch]
        ]
    }
    create_pdf(pdf_id, "static/Website Charity Application English.pdf", patient_data)

def fill_out_USMDAcharitycare_application(
        pdf_id, curr_date, guarantor_name, last_name, first_name, MI, DOS,
        hospital_acct, medical_rcrd, social_security, DOB, 
        marital_status, minors, living_with, legal_minor, patient_employed,
        spouse_employed, med_insurance, disability, disability_length,
        veteran, spouse_name, child1_name, child1_age, child2_name, child2_age,
        child3_name, child3_age, child4_name, child4_age, patient_gross,
        patient_net, spouse_gross, spouse_net, dependants_gross,
        dependants_net, pub_asst_gross, pub_asst_net, food_stamps_gross,
        food_stamps_net, soc_serc_gross, soc_serc_net, unemp_gross, unemp_net,
        strk_ben_gross, strk_ben_net, work_comp_gross, work_comp_net,
        alim_gross, alim_net, chld_sup_gross, chld_sup_net, mil_all_gross,
        mil_all_net, pen_gross, pen_net, inc_gross, inc_net, rent, utilities,
        car, groceries, credit, other_descr, other, checking, saving, CDs_IRAs,
        other_invst, properties, employer_name, spouse_employer_name,
        employer_phone, spouse_employer_phone, employer_address, spouse_employer_address,
        occupation, spouse_occupation, medicaid, county, donate, liability, assist,
        assist_identity, assist_amnt, other_info, lost_earnings, lost_time
    ):
    opt_marital_status = {
        "Married": [1.7, 3.21],
        "Single": [3, 3.21],
        "Divorced": [4.3, 3.21],
        "Widowed": [5.65, 3.21],
        "Separated": [7.15, 3.21]
    }
    opt_minors = {
        "Yes": [4.6, 3.54],
        "No": [5.6, 3.54]
    }
    opt_living_with = {
        "Yes": [4.6, 3.71],
        "No": [5.6, 3.71]
    }
    opt_legal_minor = {
        "Yes": [4.6, 3.87],
        "No": [5.6, 3.87]
    }
    opt_patient_employed = {
        "Yes": [4.6, 4.03],
        "No": [5.6, 4.03]
    }
    opt_spouse_employed = {
        "Yes": [4.6, 4.20],
        "No": [5.6, 4.20]
    }
    opt_med_insurance = {
        "Yes": [4.6, 4.36],
        "No": [5.6, 4.36]
    }
    opt_disability = {
        "Yes": [4.6, 4.52],
        "No": [5.6, 4.52]
    }
    opt_veteran = {
        "Yes": [4.6, 4.68],
        "No": [5.6, 4.68]
    }

    total_gross = int(inc_gross) + int(pen_gross) + int(alim_gross) + int(unemp_gross) + int(food_stamps_gross) + int(dependants_gross)
    total_gross += int(work_comp_gross) + int(strk_ben_gross) + int(soc_serc_gross) + int(pub_asst_gross) + int(chld_sup_gross)
    total_gross += int(patient_gross) + int(mil_all_gross) + int(spouse_gross)
    total_net = int(inc_net) + int(pen_net) + int(alim_net) + int(unemp_net) + int(food_stamps_net) + int(dependants_net)
    total_net += int(work_comp_net) + int(strk_ben_net) + int(soc_serc_net) + int(pub_asst_net) + int(chld_sup_net)
    total_net += int(patient_net) + int(mil_all_net) + int(spouse_net)
    total = int(rent) + int(utilities) + int(car) + int(groceries) + int(credit) + int(other)

    opt_medicaid = {
        "Yes": [5.95, 3.53],
        "No": [6.59, 3.53]
    }
    opt_county = {
        "Yes": [5.95, 3.70],
        "No": [6.59, 3.70]
    }
    opt_donate = {
        "Yes": [5.95, 3.85],
        "No": [6.59, 3.85]
    }
    opt_liability = {
        "Yes": [5.95, 4.21],
        "No": [6.59, 4.21]
    }
    opt_assist = {
        "Yes": [5.95, 4.37],
        "No": [6.59, 4.37]
    }

    patient_data = {
        0: [
            [curr_date, 1.5*inch, LETTER[1] - 2.43*inch],
            [guarantor_name, 5.3*inch, LETTER[1] - 2.43*inch],
            [first_name + " " + MI + " " + last_name, 2.1*inch, LETTER[1] - 2.81*inch],
            [DOS, 6.27*inch, LETTER[1] - 2.81*inch],
            [hospital_acct, 2.35*inch, LETTER[1] - 3.20*inch],
            [medical_rcrd, 6.41*inch, LETTER[1] - 3.20*inch]
        ],
        1: [
            [last_name, 2.4*inch, LETTER[1] - 2.53*inch],
            [first_name, 5*inch, LETTER[1] - 2.53*inch],
            [MI, 7.2*inch, LETTER[1] - 2.53*inch],
            [social_security, 2.1*inch, LETTER[1] - 2.86*inch],
            [DOB, 3.85*inch, LETTER[1] - 2.86*inch],
            [hospital_acct, 6.1*inch, LETTER[1] - 2.86*inch],
            
            ["X", opt_marital_status[marital_status][0]*inch, LETTER[1] - opt_marital_status[marital_status][1]*inch],
            ["X", opt_minors[minors][0]*inch, LETTER[1] - opt_minors[minors][1]*inch],
            ["X", opt_living_with[living_with][0]*inch, LETTER[1] - opt_living_with[living_with][1]*inch],
            ["X", opt_legal_minor[legal_minor][0]*inch, LETTER[1] - opt_legal_minor[legal_minor][1]*inch],
            ["X", opt_patient_employed[patient_employed][0]*inch, LETTER[1] - opt_patient_employed[patient_employed][1]*inch],
            ["X", opt_spouse_employed[spouse_employed][0]*inch, LETTER[1] - opt_spouse_employed[spouse_employed][1]*inch],
            ["X", opt_med_insurance[med_insurance][0]*inch, LETTER[1] - opt_med_insurance[med_insurance][1]*inch],
            ["X", opt_disability[disability][0]*inch, LETTER[1] - opt_disability[disability][1]*inch],
            [disability_length, 3.15*inch, LETTER[1] - 4.52*inch],
            ["X", opt_veteran[veteran][0]*inch, LETTER[1] - opt_veteran[veteran][1]*inch],

            [spouse_name, 1.65*inch, LETTER[1] - 5.15*inch],
            [child1_name, 1.4*inch, LETTER[1] - 5.35*inch],
            [child1_age, 3.75*inch, LETTER[1] - 5.35*inch],
            [child2_name, 1.4*inch, LETTER[1] - 5.52*inch],
            [child2_age, 3.75*inch, LETTER[1] - 5.52*inch],
            [child3_name, 1.4*inch, LETTER[1] - 5.69*inch],
            [child3_age, 3.75*inch, LETTER[1] - 5.69*inch],
            [child4_name, 1.4*inch, LETTER[1] - 5.86*inch],
            [child4_age, 3.75*inch, LETTER[1] - 5.86*inch],

            [patient_gross, 2.62*inch, LETTER[1] - 6.45*inch],
            [patient_net, 3.76*inch, LETTER[1] - 6.45*inch],
            [spouse_gross, 2.62*inch, LETTER[1] - 6.62*inch],
            [spouse_net, 3.76*inch, LETTER[1] - 6.62*inch],
            [dependants_gross, 2.62*inch, LETTER[1] - 6.79*inch],
            [dependants_net, 3.76*inch, LETTER[1] - 6.79*inch],
            [pub_asst_gross, 2.62*inch, LETTER[1] - 6.96*inch],
            [pub_asst_net, 3.76*inch, LETTER[1] - 6.96*inch],
            [food_stamps_gross, 2.62*inch, LETTER[1] - 7.13*inch],
            [food_stamps_net, 3.76*inch, LETTER[1] - 7.13*inch],
            [soc_serc_gross, 2.62*inch, LETTER[1] - 7.30*inch],
            [soc_serc_net, 3.76*inch, LETTER[1] - 7.30*inch],
            [unemp_gross, 2.62*inch, LETTER[1] - 7.47*inch],
            [unemp_net, 3.76*inch, LETTER[1] - 7.47*inch],
            [strk_ben_gross, 2.62*inch, LETTER[1] - 7.62*inch],
            [strk_ben_net, 3.76*inch, LETTER[1] - 7.62*inch],
            [work_comp_gross, 2.62*inch, LETTER[1] - 7.94*inch],
            [work_comp_net, 3.76*inch, LETTER[1] - 7.94*inch],
            [alim_gross, 2.62*inch, LETTER[1] - 8.11*inch],
            [alim_net, 3.76*inch, LETTER[1] - 8.11*inch],
            [chld_sup_gross, 2.62*inch, LETTER[1] - 8.28*inch],
            [chld_sup_net, 3.76*inch, LETTER[1] - 8.28*inch],
            [mil_all_gross, 2.62*inch, LETTER[1] - 8.45*inch],
            [mil_all_net, 3.76*inch, LETTER[1] - 8.45*inch],
            [pen_gross, 2.62*inch, LETTER[1] - 8.62*inch],
            [pen_net, 3.76*inch, LETTER[1] - 8.62*inch],
            [inc_gross, 2.62*inch, LETTER[1] - 9.10*inch],
            [inc_net, 3.76*inch, LETTER[1] - 9.10*inch],
            [str(total_gross), 2.62*inch, LETTER[1] - 9.34*inch],
            [str(total_net), 3.76*inch, LETTER[1] - 9.34*inch],

            [rent, 6.91*inch, LETTER[1] - 6.45*inch],
            [utilities, 6.91*inch, LETTER[1] - 6.62*inch],
            [car, 6.91*inch, LETTER[1] - 6.79*inch],
            [groceries, 6.91*inch, LETTER[1] - 6.97*inch],
            [credit, 6.91*inch, LETTER[1] - 7.13*inch],
            [other_descr, 5.07*inch, LETTER[1] - 7.46*inch],
            [other, 6.91*inch, LETTER[1] - 7.46*inch],
            [str(total), 6.91*inch, LETTER[1] - 7.94*inch],

            [checking, 4.19*inch, LETTER[1] - 9.80*inch],
            [saving, 4.19*inch, LETTER[1] - 9.97*inch],
            [CDs_IRAs, 4.19*inch, LETTER[1] - 10.14*inch],
            [other_invst, 4.19*inch, LETTER[1] - 10.31*inch],
            [properties, 4.19*inch, LETTER[1] - 10.48*inch]
        ],
        2: [
            [employer_name, 2.2*inch, LETTER[1] - 2.66*inch],
            [spouse_employer_name, 5.8*inch, LETTER[1] - 2.66*inch],
            [employer_phone, 2.2*inch, LETTER[1] - 2.85*inch],
            [spouse_employer_phone, 5.8*inch, LETTER[1] - 2.85*inch],
            [employer_address, 2.2*inch, LETTER[1] - 3.04*inch],
            [spouse_employer_address, 5.8*inch, LETTER[1] - 3.04*inch],
            [occupation, 2.2*inch, LETTER[1] - 3.24*inch],
            [spouse_occupation, 5.8*inch, LETTER[1] - 3.24*inch],

            ["X", opt_medicaid[medicaid][0]*inch, LETTER[1] - opt_medicaid[medicaid][1]*inch],
            ["X", opt_county[county][0]*inch, LETTER[1] - opt_county[county][1]*inch],
            ["X", opt_donate[donate][0]*inch, LETTER[1] - opt_donate[donate][1]*inch],
            ["X", opt_liability[liability][0]*inch, LETTER[1] - opt_liability[liability][1]*inch],
            ["X", opt_assist[assist][0]*inch, LETTER[1] - opt_assist[assist][1]*inch],
            [assist_identity, 5.8*inch, LETTER[1] - 4.53*inch],
            [assist_amnt, 5.8*inch, LETTER[1] - 4.68*inch],

            [other_info, 0.85*inch, LETTER[1] - 5.30*inch],

            [lost_earnings, 5.99*inch, LETTER[1] - 6.05*inch],
            [lost_time, 5.9*inch, LETTER[1] - 6.33*inch]
        ]
    }
    create_pdf(pdf_id, "static/Website Charity Application USMD Arlington English.pdf", patient_data)


app = Flask(__name__)
secret_key = ""
for i in range(3):
    secret_key += str(random.randint(0,9))
for i in range(7):
    secret_key += random.choice(string.ascii_letters)
for i in range(2):
    secret_key += str(random.randint(0,9))
app.secret_key = secret_key


def create_pdf_id():
    current_time = str(datetime.datetime.now().strftime("%H_%M_%S"))
    remote_address = str(request.remote_addr.replace(".", "_"))
    random_number = str(random.randint(10000, 99999))
    pdf_id = "{}__{}__{}.pdf".format(current_time, remote_address, random_number)
    return pdf_id

def create_session():
    if 'pdf_id' not in session:
        session['pdf_id'] = create_pdf_id()
    if 'household_size' not in session:
        session['household_size'] = 1
    if 'estimated_annual_income' not in session:
        session['estimated_annual_income'] = 0
    if 'balance_due' not in session:
        session['balance_due'] = 0
    if 'estimated_cost' not in session:
        session['estimated_cost'] = 0

def create_blank_session():
    if 'pdf_id' not in session:
        session['pdf_id'] = create_pdf_id()
    session['household_size'] = 1
    session['estimated_annual_income'] = 0
    session['balance_due'] = 0
    session['estimated_cost'] = 0

def clear_session():
    [session.pop(key) for key in list(session.keys())]
    session.clear()
    create_blank_session()


@app.route("/")
def redirect_to_home():
    return redirect("/home")

@app.route("/home")
def home():
    create_blank_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template("home.html")

@app.route("/about")
def display_about():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template("about.html")

@app.route("/charitycare")
def display_charity_care():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template("charity_care.html")

@app.route("/financiallytest")
def financially_indignant_form():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template(
        "financially_indignant_form.html",
        household_size = session['household_size'],
        estimated_annual_income = session['estimated_annual_income']
    )

@app.route("/financiallyresult", methods = ["POST", "GET"])
def financially_indignant_result():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    if request.method == "POST":
        session['household_size'] = household_size = int(request.form["household_size"])
        session['estimated_annual_income'] = estimated_annual_income = int(request.form["estimated_annual_income"])
        if determine_financially_indignant(household_size, estimated_annual_income) != 0:
            return render_template(
                "qualify_success_results.html",
                qualification_type = "Financially",
                pct_discount=100
            )
        else:
            return redirect("/medicallytest")
    else:
        return redirect("/home")

@app.route("/locationsearch")
def select_location():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template(
        "location_search.html"
    )

@app.route("/chargemaster", methods = ["POST", "GET"])
def select_procedure():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    if request.method == "GET":
        return redirect("/locationsearch")
    location = request.form["facility"]
    return render_template(
        "chargemaster.html",
        procedures = get_chargemaster(location)
    )

@app.route("/revealcost", methods = ["POST", "GET"])
def reveal_cost():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    if request.method == "GET":
        return redirect("/locationsearch")
    session['estimated_cost'] = int(float(request.form["cost"]))
    return render_template(
        "reveal_cost.html",
        cost = "{:,}".format(session['estimated_cost'])
    )

@app.route("/medicallytest", methods=["POST", "GET"])
def medically_indignant_form():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    if session['estimated_cost'] != 0:
        return render_template(
            "medically_indignant_form.html",
            household_size = session['household_size'],
            estimated_annual_income = session['estimated_annual_income'],
            balance_due = session['estimated_cost'],
            is_estimate = " (Estimate)"
        )
    else:
        return render_template(
            "medically_indignant_form.html",
            household_size = session['household_size'],
            estimated_annual_income = session['estimated_annual_income'],
            balance_due = session['estimated_cost'],
            is_estimate = ""
    )

@app.route("/medicallyresult", methods = ["POST", "GET"])
def medically_indignant_result():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    if request.method == "POST":
        session['household_size'] = household_size = int(request.form["household_size"])
        session['estimated_annual_income'] = estimated_annual_income = int(request.form["estimated_annual_income"])
        session['balance_due'] = balance_due = int(request.form["balance_due"])
        if determine_financially_indignant(household_size, estimated_annual_income) != 0:
            return render_template(
                "qualify_success_results.html",
                qualification_type = "Financially",
                pct_discount=100
            )
        if determine_medically_indignant(household_size, estimated_annual_income, balance_due) != 0:
            if determine_catastrophic_medically_indignant(household_size, estimated_annual_income, balance_due) != 0:
                return render_template(
                    "qualify_success_results.html",
                    qualification_type = "Catastrophic Medically",
                    pct_discount=100*determine_catastrophic_medically_indignant(household_size, estimated_annual_income, balance_due)
                )
            return render_template(
                    "qualify_success_results.html",
                    qualification_type = "Medically",
                    pct_discount=100*determine_medically_indignant(household_size, estimated_annual_income, balance_due)
                )
        elif determine_tier2_medically_indignant(household_size, estimated_annual_income, balance_due) != 0:
            return render_template(
                    "qualify_success_results.html",
                    qualification_type = "Tier 2 Medically",
                    pct_discount=100*determine_tier2_medically_indignant(household_size, estimated_annual_income, balance_due)
                )
        elif determine_catastrophic_medically_indignant(household_size, estimated_annual_income, balance_due) != 0:
            return render_template(
                "qualify_success_results.html",
                qualification_type = "Catastrophic Medically",
                pct_discount=100*determine_catastrophic_medically_indignant(household_size, estimated_annual_income, balance_due)
            )
        else:
            return render_template("qualify_failure_results.html")
    else:
        return redirect("/home")

@app.route("/blankapplication")
def blank_application():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template("blank_application.html")

@app.route("/blankapplicationTHP")
def blank_application_THP():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template("blank_application_THP.html")

@app.route("/blankapplicationUSMDA")
def blank_application_USMDA():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template("blank_application_USMDA.html")

@app.route("/completeapplication")
def complete_application():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    return render_template("application_completer.html")

@app.route("/completedapplication", methods = ["POST", "GET"])
def display_completed_application():
    create_session()
    delete_file("static/{}".format(session['pdf_id']))
    if request.method == "POST":
        application_data = {}
        for key, value in request.form.items():
            application_data[key] = str(value)
        if application_data["facility"] == "USMD Hospital at Arlington":
            fill_out_USMDAcharitycare_application(
                "static/" + session["pdf_id"], application_data["curr_date"], application_data["guarantor_name"],
                application_data["last_name"], application_data["first_name"], application_data["MI"], application_data["DOS"],
                application_data["hospital_acct"], application_data['medical_rcrd'], application_data['social_security'],
                application_data["DOB"], application_data['marital_status'], application_data['minors'],
                application_data["living_with"], application_data['legal_minor'], application_data['patient_employed'],
                application_data["spouse_employed"], application_data['med_insurance'], application_data['disability'],
                application_data["disability_length"], application_data['veteran'], application_data['spouse_name'],
                application_data["child1_name"], application_data['child1_age'], application_data['child2_name'],
                application_data["child2_age"], application_data['child3_name'], application_data['child3_age'],
                application_data["child4_name"], application_data['child4_age'], application_data['patient_gross'],
                application_data["patient_net"], application_data['spouse_gross'], application_data['spouse_net'],
                application_data["dependants_gross"], application_data['dependants_net'], application_data['pub_asst_gross'],
                application_data["pub_asst_net"], application_data['food_stamps_gross'], application_data['food_stamps_net'],
                application_data["soc_serc_gross"], application_data['soc_serc_net'], application_data['unemp_gross'],
                application_data["unemp_net"], application_data['strk_ben_gross'], application_data['strk_ben_net'],
                application_data["work_comp_gross"], application_data['work_comp_net'], application_data['alim_gross'],
                application_data["alim_net"], application_data['chld_sup_gross'], application_data['chld_sup_net'],
                application_data["mil_all_gross"], application_data['mil_all_net'], application_data['pen_gross'],
                application_data["pen_net"], application_data['inc_gross'], application_data['inc_net'],
                application_data["rent"], application_data['utilities'], application_data['car'], application_data['groceries'],
                application_data["credit"], application_data['other_descr'], application_data['other'],
                application_data["checking"], application_data['saving'], application_data['CDs_IRAs'],
                application_data["other_invst"], application_data["properties"], application_data["employer_name"],
                application_data["spouse_employer_name"], application_data["employer_phone"],
                application_data["spouse_employer_phone"], application_data['employer_address'],
                application_data["spouse_employer_address"], application_data["occupation"],
                application_data["spouse_occupation"], application_data["medicaid"], application_data["county"],
                application_data["donate"], application_data["liability"], application_data["assist"],
                application_data["assist_identity"], application_data["assist_amnt"], application_data["other_info"],
                application_data["lost_earnings"], application_data["lost_time"]
            )
        elif (
            application_data["facility"] == "Texas Health Center for Diagnostics & Surgery Plano"
            or application_data["facility"] == "Texas Health Harris Methodist Southlake"
            or application_data["facility"] == "Texas Health Presbyterian Hospital Flower Mound"
            or application_data["facility"] == "Texas Health Presbyterian Hospital Rockwall"
            or application_data["facility"] == "Texas Institute for Surgery at Texas Health Presbyterian Dallas"
        ):
            fill_out_THP_application(
                "static/" + session["pdf_id"], application_data["curr_date"], application_data["guarantor_name"],
                application_data["last_name"], application_data["first_name"], application_data["MI"], application_data["DOS"],
                application_data["hospital_acct"], application_data['medical_rcrd'], application_data['facility'], application_data['social_security'],
                application_data["DOB"], application_data['marital_status'], application_data['minors'],
                application_data["living_with"], application_data['legal_minor'], application_data['patient_employed'],
                application_data["spouse_employed"], application_data['med_insurance'], application_data['disability'],
                application_data["disability_length"], application_data['veteran'], application_data['spouse_name'],
                application_data["child1_name"], application_data['child1_age'], application_data['child2_name'],
                application_data["child2_age"], application_data['child3_name'], application_data['child3_age'],
                application_data["child4_name"], application_data['child4_age'], application_data['patient_gross'],
                application_data["patient_net"], application_data['spouse_gross'], application_data['spouse_net'],
                application_data["dependants_gross"], application_data['dependants_net'], application_data['pub_asst_gross'],
                application_data["pub_asst_net"], application_data['food_stamps_gross'], application_data['food_stamps_net'],
                application_data["soc_serc_gross"], application_data['soc_serc_net'], application_data['unemp_gross'],
                application_data["unemp_net"], application_data['strk_ben_gross'], application_data['strk_ben_net'],
                application_data["work_comp_gross"], application_data['work_comp_net'], application_data['alim_gross'],
                application_data["alim_net"], application_data['chld_sup_gross'], application_data['chld_sup_net'],
                application_data["mil_all_gross"], application_data['mil_all_net'], application_data['pen_gross'],
                application_data["pen_net"], application_data['inc_gross'], application_data['inc_net'],
                application_data["rent"], application_data['utilities'], application_data['car'], application_data['groceries'],
                application_data["credit"], application_data['other_descr'], application_data['other'],
                application_data["checking"], application_data['saving'], application_data['CDs_IRAs'],
                application_data["other_invst"], application_data["properties"], application_data["employer_name"],
                application_data["spouse_employer_name"], application_data["employer_phone"],
                application_data["spouse_employer_phone"], application_data['employer_address'],
                application_data["spouse_employer_address"], application_data["occupation"],
                application_data["spouse_occupation"], application_data["medicaid"], application_data["county"],
                application_data["donate"], application_data["liability"], application_data["assist"],
                application_data["assist_identity"], application_data["assist_amnt"], application_data["other_info"],
                application_data["lost_earnings"], application_data["lost_time"]
            )
        else:
            fill_out_charitycare_application(
                "static/" + session["pdf_id"], application_data["curr_date"], application_data["guarantor_name"],
                application_data["last_name"], application_data["first_name"], application_data["MI"], application_data["DOS"],
                application_data["hospital_acct"], application_data['medical_rcrd'], application_data['facility'], application_data['social_security'],
                application_data["DOB"], application_data['marital_status'], application_data['minors'],
                application_data["living_with"], application_data['legal_minor'], application_data['patient_employed'],
                application_data["spouse_employed"], application_data['med_insurance'], application_data['disability'],
                application_data["disability_length"], application_data['veteran'], application_data['spouse_name'],
                application_data["child1_name"], application_data['child1_age'], application_data['child2_name'],
                application_data["child2_age"], application_data['child3_name'], application_data['child3_age'],
                application_data["child4_name"], application_data['child4_age'], application_data['patient_gross'],
                application_data["patient_net"], application_data['spouse_gross'], application_data['spouse_net'],
                application_data["dependants_gross"], application_data['dependants_net'], application_data['pub_asst_gross'],
                application_data["pub_asst_net"], application_data['food_stamps_gross'], application_data['food_stamps_net'],
                application_data["soc_serc_gross"], application_data['soc_serc_net'], application_data['unemp_gross'],
                application_data["unemp_net"], application_data['strk_ben_gross'], application_data['strk_ben_net'],
                application_data["work_comp_gross"], application_data['work_comp_net'], application_data['alim_gross'],
                application_data["alim_net"], application_data['chld_sup_gross'], application_data['chld_sup_net'],
                application_data["mil_all_gross"], application_data['mil_all_net'], application_data['pen_gross'],
                application_data["pen_net"], application_data['inc_gross'], application_data['inc_net'],
                application_data["rent"], application_data['utilities'], application_data['car'], application_data['groceries'],
                application_data["credit"], application_data['other_descr'], application_data['other'],
                application_data["checking"], application_data['saving'], application_data['CDs_IRAs'],
                application_data["other_invst"], application_data["properties"], application_data["employer_name"],
                application_data["spouse_employer_name"], application_data["employer_phone"],
                application_data["spouse_employer_phone"], application_data['employer_address'],
                application_data["spouse_employer_address"], application_data["occupation"],
                application_data["spouse_occupation"], application_data["medicaid"], application_data["county"],
                application_data["donate"], application_data["liability"], application_data["assist"],
                application_data["assist_identity"], application_data["assist_amnt"], application_data["other_info"],
                application_data["lost_earnings"], application_data["lost_time"]
            )
        return render_template(
            "completed_application.html",
            pdf_src = session['pdf_id']
        )
    else:
        return redirect("/home")

@app.route("/finished")
def final_page():
    delete_file("static/{}".format(session['pdf_id']))
    clear_session()
    return render_template("finished.html")

if __name__ == "__main__":
    app.run()